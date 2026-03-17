# --- IMPORTACIONES: Traemos las herramientas que necesitamos ---
# 'render': Sirve para cargar un archivo HTML y mostrarlo al usuario.
# 'redirect': Sirve para enviar al usuario a otra página.
# 'get_object_or_404': Busca algo en la base de datos, y si no existe, da error 404.
from django.shortcuts import render, redirect, get_object_or_404

# 'login_required': Es un guardia. Si no estás logueado, te manda al login automáticamente.
from django.contrib.auth.decorators import login_required
# Importa el sistema de mensajes de Django para mostrar notificaciones al usuario
from django.contrib import messages
# 'JsonResponse': Para devolver datos en formato JSON (útil para APIs y JavaScript).
from django.http import JsonResponse
# Importa pandas para leer y manipular archivos Excel
import pandas as pd
# Importa Decimal para manejar operaciones monetarias con precisión decimal
from decimal import Decimal
# 'csrf_exempt': Desactiva una protección de seguridad de Django. 
# (Nota: Se usa aquí para simplificar llamadas desde JS, pero en producción se maneja de otra forma).
from django.views.decorators.csrf import csrf_exempt

# 'models': Nos ayuda a hacer cálculos en la base de datos (como sumas).
from django.db import models

# Importa datetime para trabajar con fechas
from datetime import datetime

# Importaciones para exportar Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse

# Importamos nuestras tablas (Modelos) definidas en models.py de esta app.
from .models import Transaccion, Deuda, Perfil

# Importamos el modelo de Usuario de Django.
from django.contrib.auth.models import User

# 'json': Para poder leer datos que nos envían en formato texto (crudo) y convertirlos a Python.
import json

# 'Decimal': Para manejar dinero con precisión matemática (evita errores de los números flotantes).
from decimal import Decimal
from django.db.models import Sum

# --- VISTA PRINCIPAL: DASHBOARD ---
# El símbolo '@' es un "decorador". Modifica la función de abajo.
# Aquí dice: "Solo ejecuta esta función si el usuario ha iniciado sesión".

@login_required  # ← Este decorador es OBLIGATORIO, faltaba en la segunda versión
def dashboard(request):
    # Leer filtros del GET
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    
    # Queryset base — SOLO del usuario logueado
    transacciones = Transaccion.objects.filter(usuario=request.user)
    
    # Aplicar filtros si existen
    if anio:
        transacciones = transacciones.filter(fecha__year=anio)
    if mes:
        transacciones = transacciones.filter(fecha__month=mes)
    
    # Calcular totales CON los filtros aplicados
    total_ingresos = transacciones.filter(tipo='INGRESO').aggregate(
        total=Sum('monto'))['total'] or 0
    total_gastos = transacciones.filter(tipo='GASTO').aggregate(
        total=Sum('monto'))['total'] or 0
    balance = total_ingresos - total_gastos
    
    # Últimas transacciones filtradas
    ultimas_transacciones = transacciones.order_by('-fecha')[:10]
    
    # Datos para gráfica
    categorias_data = transacciones.filter(tipo='GASTO').values(
        'categoria').annotate(total=Sum('monto')).order_by('-total')
    
    categorias_json = json.dumps([c['categoria'] for c in categorias_data])
    valores_json = json.dumps([float(c['total']) for c in categorias_data])
    
    return render(request, 'base/dashboard.html', {
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'balance': balance,
        'ultimas_transacciones': ultimas_transacciones,
        'categorias_json': categorias_json,
        'valores_json': valores_json,
    })


# --- VISTA DE LISTA DE TRANSACCIONES ---
@login_required
def transacciones_lista(request):
    anio = request.GET.get('anio', '')
    mes  = request.GET.get('mes', '')
    tipo = request.GET.get('tipo', '')
    categoria = request.GET.get('categoria', '')

    transacciones = Transaccion.objects.filter(usuario=request.user).order_by('-fecha')

    if anio:
        transacciones = transacciones.filter(fecha__year=anio)
    if mes:
        transacciones = transacciones.filter(fecha__month=mes)
    if tipo:
        transacciones = transacciones.filter(tipo=tipo)
    if categoria:
        transacciones = transacciones.filter(categoria__icontains=categoria)

    categorias_disponibles = Transaccion.objects.filter(
        usuario=request.user
    ).values_list('categoria', flat=True).distinct().order_by('categoria')

    # --- DATOS PARA GRÁFICAS ---
    base_qs = Transaccion.objects.filter(usuario=request.user)
    if anio:
        base_qs = base_qs.filter(fecha__year=anio)
    if mes:
        base_qs = base_qs.filter(fecha__month=mes)

    total_ingresos = base_qs.filter(tipo='INGRESO').aggregate(t=Sum('monto'))['t'] or 0
    total_gastos   = base_qs.filter(tipo='GASTO').aggregate(t=Sum('monto'))['t'] or 0

    top_gastos = base_qs.filter(tipo='GASTO').values('categoria').annotate(
        total=Sum('monto')).order_by('-total')[:8]

    context = {
        'transacciones': transacciones,
        'categorias_disponibles': categorias_disponibles,
        'total_ingresos': float(total_ingresos),
        'total_gastos': float(total_gastos),
        'top_categorias_json': json.dumps([x['categoria'] for x in top_gastos]),
        'top_valores_json':    json.dumps([float(x['total']) for x in top_gastos]),
    }
    return render(request, 'base/transacciones.html', context)



@login_required
def deudas_lista(request):
    deudas = Deuda.objects.filter(usuario=request.user, activa=True).order_by('-fecha_creacion')
    
    # Datos para gráficas del dashboard
    total_deuda = sum(float(d.saldo_pendiente()) for d in deudas)
    total_pagado = sum(float(d.total_pagado()) for d in deudas)
    
    # Preparar datos de cada deuda para el template
    deudas_data = []
    for d in deudas:
        deudas_data.append({
            'deuda': d,
            'total_pagado': float(d.total_pagado()),
            'saldo_pendiente': float(d.saldo_pendiente()),
            'meses_pagados': d.meses_pagados(),
            'meses_restantes': d.meses_restantes(),
            'porcentaje_pagado': d.porcentaje_pagado(),
            'interes_pagado': d.interes_pagado(),
            'interes_total': float(d.interes_total()),
        })
    
    # JSON para gráficas
    nombres_json = json.dumps([d['deuda'].nombre for d in deudas_data])
    pendientes_json = json.dumps([d['saldo_pendiente'] for d in deudas_data])
    pagados_json = json.dumps([d['total_pagado'] for d in deudas_data])

    return render(request, 'base/deudas.html', {
        'deudas_data': deudas_data,
        'total_deuda': total_deuda,
        'total_pagado': total_pagado,
        'nombres_json': nombres_json,
        'pendientes_json': pendientes_json,
        'pagados_json': pagados_json,
    })


@login_required
@csrf_exempt
def crear_deuda(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            deuda = Deuda.objects.create(
                usuario=request.user,
                nombre=data['nombre'],
                tipo=data['tipo'],
                valor_inicial=Decimal(str(data['valor_inicial'])),
                valor_total_a_pagar=Decimal(str(data['valor_total_a_pagar'])),
                cuota_mensual=Decimal(str(data['cuota_mensual'])),
                total_meses=int(data['total_meses']),
                fecha_inicio=data['fecha_inicio'],
                descripcion=data.get('descripcion', '')
            )
            return JsonResponse({'success': True, 'id': deuda.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})


@login_required
@csrf_exempt
def pagar_deuda(request, id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            deuda = Deuda.objects.get(id=id, usuario=request.user)
            monto = Decimal(str(data['monto']))
            fecha_str = data.get('fecha')
            
            from django.utils import timezone
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
            
            # Crear transacción de gasto asociada a la deuda
            Transaccion.objects.create(
                usuario=request.user,
                tipo='GASTO',
                categoria='deuda' if deuda.tipo == 'DEUDA' else 'deuda-reportado',
                monto=monto,
                fecha=fecha,
                descripcion=f'Pago deuda: {deuda.nombre}',
                deuda=deuda
            )
            
            # Si el saldo queda en 0, marcar como inactiva
            if deuda.saldo_pendiente() <= 0:
                deuda.activa = False
                deuda.save()
            
            return JsonResponse({
                'success': True,
                'saldo_pendiente': float(deuda.saldo_pendiente()),
                'porcentaje': deuda.porcentaje_pagado()
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})


@login_required
@csrf_exempt  
def eliminar_deuda(request, id):
    if request.method == 'POST':
        try:
            deuda = Deuda.objects.get(id=id, usuario=request.user)
            deuda.activa = False  # Soft delete
            deuda.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})

# --- API PARA CREAR TRANSACCIÓN ---
@login_required
@csrf_exempt
def crear_transaccion(request):
    if request.method == 'POST':
        # Convertimos los datos enviados a Python.
        data = json.loads(request.body)

        # Obtener fecha del JSON o usar hoy
        fecha_str = data.get('fecha')
        if fecha_str:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            from django.utils import timezone
            fecha = timezone.now().date()

        # Crea la transacción en la base de datos.
        # Decimal(...) convierte el texto del monto a un tipo numérico preciso para dinero.
        transaccion = Transaccion.objects.create(
            usuario=request.user,
            tipo=data['tipo'],
            categoria=data['categoria'],
            monto=Decimal(data['monto']),
            # .get(..., '') intenta buscar descripcion, si no existe, pone un string vacío.
            descripcion=data.get('descripcion', ''),
            fecha=fecha,
        )

        # Responde al JavaScript que todo salió bien.
        return JsonResponse({
            'success': True,
            'id': transaccion.id
        })

    return JsonResponse({'success': False})


# --- VISTA PÚBLICA (HOME) ---
def index(request):
    # Esta función no tiene @login_required, así que cualquiera puede verla.
    # Simplemente muestra la página de inicio (Landing Page).
    return render(request, 'base/index.html')


@login_required
@csrf_exempt
def eliminar_transaccion(request, id):
    if request.method == 'POST':
        try:
            transaccion = Transaccion.objects.get(id=id, usuario=request.user)
            transaccion.delete()
            return JsonResponse({'success': True})
        except Transaccion.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transacción no encontrada'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ── AGREGAR en views.py, junto a las otras vistas de deuda ──

@login_required
@csrf_exempt
def editar_deuda(request, id):
    """Editar los datos de una deuda existente"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            deuda = Deuda.objects.get(id=id, usuario=request.user)

            deuda.nombre              = data.get('nombre', deuda.nombre)
            deuda.tipo                = data.get('tipo', deuda.tipo)
            deuda.valor_inicial       = Decimal(str(data.get('valor_inicial', deuda.valor_inicial)))
            deuda.valor_total_a_pagar = Decimal(str(data.get('valor_total_a_pagar', deuda.valor_total_a_pagar)))
            deuda.cuota_mensual       = Decimal(str(data.get('cuota_mensual', deuda.cuota_mensual)))
            deuda.total_meses         = int(data.get('total_meses', deuda.total_meses))
            deuda.fecha_inicio        = data.get('fecha_inicio', deuda.fecha_inicio)
            deuda.descripcion         = data.get('descripcion', deuda.descripcion)
            deuda.save()

            return JsonResponse({'success': True})
        except Deuda.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Deuda no encontrada'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@csrf_exempt
def obtener_transaccion(request, id):
    """Obtener datos de una transacción para ver/editar"""
    try:
        transaccion = Transaccion.objects.get(id=id, usuario=request.user)
        return JsonResponse({
            'success': True,
            'transaccion': {
                'id': transaccion.id,
                'fecha': transaccion.fecha.strftime('%d/%m/%Y'),
                'tipo': transaccion.tipo,
                'categoria': transaccion.categoria,
                'monto': float(transaccion.monto),
                'descripcion': transaccion.descripcion or '',
            }
        })
    except Transaccion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transacción no encontrada'})


@login_required
@csrf_exempt
def editar_transaccion(request, id):
    """Editar una transacción existente"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            transaccion = Transaccion.objects.get(id=id, usuario=request.user)
            
            # Actualizar campos permitidos
            transaccion.categoria = data.get('categoria', transaccion.categoria)
            transaccion.monto = Decimal(data.get('monto', transaccion.monto))
            transaccion.descripcion = data.get('descripcion', transaccion.descripcion)
            transaccion.save()
            
            return JsonResponse({'success': True})
        except Transaccion.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transacción no encontrada'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required  # Decorador: solo usuarios logueados pueden acceder
def importar_excel(request):
    """
    Vista para importar transacciones desde un archivo Excel
    """
    # Solo procesamos si es POST (formulario enviado) y hay archivo subido
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            # Obtenemos el archivo del formulario
            archivo = request.FILES['archivo_excel']
            
            # Pandas lee el Excel y lo convierte en DataFrame (tabla)
            df = pd.read_excel(archivo, dtype = {'fecha': str})
            # Leer pestaña específica TRANSACCION de archivo .xlsb
           # df = pd.read_excel(archivo, sheet_name='TRANSACCION', engine='pyxlsb')
            
            # Definimos qué columnas debe tener el Excel
            columnas_requeridas = ['tipo', 'categoria', 'monto', 'descripcion', 'fecha']
            
            # Verificamos que el Excel tenga todas las columnas necesarias
            if not all(col in df.columns for col in columnas_requeridas):
                messages.error(request, f'El Excel debe tener estas columnas: {", ".join(columnas_requeridas)}')
                return redirect('base:transacciones')
            
            # Contadores para el resumen final
            creadas = 0      # Transacciones exitosas
            errores = []     # Lista de errores encontrados
            
            # iterrows() recorre el DataFrame fila por fila
            # index = número de fila (empieza en 0), row = datos de la fila
            for index, row in df.iterrows():
                try:
                    # Limpiamos el texto: mayúsculas y sin espacios
                    tipo = str(row['tipo']).upper().strip()
                    
                    # Validamos que el tipo sea correcto
                    if tipo not in ['INGRESO', 'GASTO']:
                        errores.append(f"Fila {index+2}: Tipo debe ser INGRESO o GASTO")
                        continue  # Saltamos a la siguiente fila
                    # Convertir fecha
                    fecha = pd.to_datetime(row['fecha'], errors='coerce').date()
                    if pd.isna(fecha):
                        fecha = datetime.now().date()
                    # Validar que la fecha no sea futura (opcional)
                    if fecha > datetime.now().date():
                        errores.append(f"Fila {index+2}: La fecha {fecha} es futura")
                        continue

                   
                    # Creamos el objeto en la base de datos
                    Transaccion.objects.create(
                        usuario=request.user,  # El usuario logueado
                        tipo=tipo,
                        categoria=str(row['categoria']),
                        monto=Decimal(str(row['monto'])),  # Decimal para precisión monetaria
                        fecha=fecha,
                        descripcion=str(row['descripcion']) if pd.notna(row['descripcion']) else ''
                    )
                    creadas += 1  # Sumamos al contador de éxitos
                    
                except Exception as e:
                    # Capturamos cualquier error de esta fila específica
                    errores.append(f"Fila {index+2}: {str(e)}")
            
            # Mostramos mensajes de resultado al usuario
            if creadas > 0:
                messages.success(request, f'✅ Se importaron {creadas} transacciones correctamente')
            if errores:
                for error in errores[:5]:  # Solo mostramos los primeros 5 errores
                    messages.warning(request, error)
            
            return redirect('base:transacciones')
            
        except Exception as e:
            # Error general (ej: archivo corrupto, no es Excel)
            messages.error(request, f'❌ Error al procesar el archivo: {str(e)}')
            return redirect('base:transacciones')
    
    # Si no es POST o no hay archivo, redirigimos
    return redirect('base:transacciones')


# ============================================
# VISTAS DE REGISTRO Y GESTIÓN DE USUARIOS
# ============================================

def registro(request):
    """
    Vista para registrar nuevos usuarios.
    Crea un usuario de Django y un perfil asociado.
    """
    # Si el usuario ya está autenticado, redirigir al dashboard
    if request.user.is_authenticated:
        return redirect('base:dashboard')

    if request.method == 'POST':
        # Obtener datos del formulario
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        telefono = request.POST.get('telefono', '').strip()

        # Validaciones
        errores = []

        if not username:
            errores.append('El nombre de usuario es obligatorio.')
        elif len(username) < 3:
            errores.append('El nombre de usuario debe tener al menos 3 caracteres.')
        elif User.objects.filter(username=username).exists():
            errores.append('Este nombre de usuario ya está en uso.')

        if not email:
            errores.append('El correo electrónico es obligatorio.')
        elif User.objects.filter(email=email).exists():
            errores.append('Este correo electrónico ya está registrado.')

        if not password:
            errores.append('La contraseña es obligatoria.')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres.')

        if password != password_confirm:
            errores.append('Las contraseñas no coinciden.')

        # Si hay errores, mostrarlos
        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'base/registro.html', {
                'username': username,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'telefono': telefono,
            })

        # Crear el usuario
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )

            # Crear el perfil asociado
            Perfil.objects.create(
                usuario=user,
                telefono=telefono
            )

            messages.success(request, '¡Cuenta creada exitosamente! Ahora puedes iniciar sesión.')
            return redirect('login')

        except Exception as e:
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
            return render(request, 'base/registro.html', {
                'username': username,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'telefono': telefono,
            })

    # GET: Mostrar formulario vacío
    return render(request, 'base/registro.html')


@login_required
def mi_perfil(request):
    """
    Vista para ver y editar el perfil del usuario logueado.
    """
    # Obtener o crear el perfil del usuario
    perfil, creado = Perfil.objects.get_or_create(
        usuario=request.user,
        defaults={'avatar': '👤'}
    )

    if request.method == 'POST':
        # Actualizar datos del usuario
        request.user.first_name = request.POST.get('first_name', '').strip()
        request.user.last_name = request.POST.get('last_name', '').strip()
        request.user.email = request.POST.get('email', '').strip()
        request.user.save()

        # Actualizar datos del perfil
        perfil.telefono = request.POST.get('telefono', '').strip()
        perfil.direccion = request.POST.get('direccion', '').strip()
        perfil.ciudad = request.POST.get('ciudad', '').strip()
        perfil.pais = request.POST.get('pais', '').strip()
        perfil.bio = request.POST.get('bio', '').strip()
        perfil.avatar = request.POST.get('avatar', '👤').strip()

        # Fecha de nacimiento
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
        if fecha_nacimiento:
            try:
                perfil.fecha_nacimiento = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
            except ValueError:
                pass

        perfil.save()

        messages.success(request, 'Perfil actualizado correctamente.')
        return redirect('base:mi_perfil')

    return render(request, 'base/perfil.html', {
        'perfil': perfil,
    })


@login_required
def cambiar_password(request):
    """
    Vista para cambiar la contraseña del usuario.
    """
    if request.method == 'POST':
        password_actual = request.POST.get('password_actual', '')
        password_nueva = request.POST.get('password_nueva', '')
        password_confirmar = request.POST.get('password_confirmar', '')

        errores = []

        # Verificar contraseña actual
        if not request.user.check_password(password_actual):
            errores.append('La contraseña actual es incorrecta.')

        # Validar nueva contraseña
        if len(password_nueva) < 6:
            errores.append('La nueva contraseña debe tener al menos 6 caracteres.')

        if password_nueva != password_confirmar:
            errores.append('Las contraseñas no coinciden.')

        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            request.user.set_password(password_nueva)
            request.user.save()
            messages.success(request, 'Contraseña actualizada correctamente. Por favor, vuelve a iniciar sesión.')
            return redirect('login')

    return render(request, 'base/cambiar_password.html')


# ============================================
# VISTA PARA EXPORTAR TRANSACCIONES A EXCEL CON DASHBOARD
# ============================================

@login_required
def exportar_excel(request):
    """
    Vista para exportar transacciones a Excel con dashboard incluido.
    Crea un archivo Excel con:
    - Hoja de datos de transacciones
    - Dashboard con resumen y gráficos
    """
    # Obtener filtros de la URL
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    tipo = request.GET.get('tipo', '')
    categoria = request.GET.get('categoria', '')

    # Queryset base - solo del usuario logueado
    transacciones = Transaccion.objects.filter(usuario=request.user).order_by('-fecha')

    # Aplicar filtros si existen
    if anio:
        transacciones = transacciones.filter(fecha__year=anio)
    if mes:
        transacciones = transacciones.filter(fecha__month=mes)
    if tipo:
        transacciones = transacciones.filter(tipo=tipo)
    if categoria:
        transacciones = transacciones.filter(categoria__icontains=categoria)

    # Calcular totales
    total_ingresos = transacciones.filter(tipo='INGRESO').aggregate(
        total=Sum('monto'))['total'] or 0
    total_gastos = transacciones.filter(tipo='GASTO').aggregate(
        total=Sum('monto'))['total'] or 0
    balance = total_ingresos - total_gastos

    # Datos para gráficos - Top categorías de gasto
    categorias_data = transacciones.filter(tipo='GASTO').values(
        'categoria').annotate(total=Sum('monto')).order_by('-total')[:10]

    # Crear el libro de Excel
    wb = Workbook()

    # ============================================
    # HOJA 1: DASHBOARD
    # ============================================
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    title_font = Font(bold=True, size=16, color="333333")
    subtitle_font = Font(bold=True, size=12, color="667eea")
    value_font = Font(bold=True, size=18, color="333333")
    ingreso_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    gasto_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    balance_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título del Dashboard
    ws_dashboard['A1'] = "DASHBOARD DE TRANSACCIONES"
    ws_dashboard['A1'].font = Font(bold=True, size=20, color="667eea")
    ws_dashboard.merge_cells('A1:F1')
    ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Información del usuario y fecha
    ws_dashboard['A2'] = f"Usuario: {request.user.username}"
    ws_dashboard['B2'] = f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_dashboard['A3'] = f"Filtros aplicados: Año={anio or 'Todos'}, Mes={mes or 'Todos'}, Tipo={tipo or 'Todos'}, Categoría={categoria or 'Todas'}"
    ws_dashboard.merge_cells('A3:F3')

    # Espacio
    current_row = 5

    # ============================================
    # RESUMEN GENERAL
    # ============================================
    ws_dashboard[f'A{current_row}'] = "RESUMEN GENERAL"
    ws_dashboard[f'A{current_row}'].font = subtitle_font
    ws_dashboard.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2

    # Tarjetas de resumen
    # Ingresos
    ws_dashboard[f'A{current_row}'] = "TOTAL INGRESOS"
    ws_dashboard[f'A{current_row}'].font = Font(bold=True, size=11, color="065f46")
    ws_dashboard[f'A{current_row}'].fill = ingreso_fill
    ws_dashboard[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'A{current_row}'].border = thin_border

    ws_dashboard[f'B{current_row}'] = float(total_ingresos)
    ws_dashboard[f'B{current_row}'].font = Font(bold=True, size=16, color="065f46")
    ws_dashboard[f'B{current_row}'].fill = ingreso_fill
    ws_dashboard[f'B{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'B{current_row}'].number_format = '$#,##0.00'
    ws_dashboard[f'B{current_row}'].border = thin_border

    # Gastos
    ws_dashboard[f'D{current_row}'] = "TOTAL GASTOS"
    ws_dashboard[f'D{current_row}'].font = Font(bold=True, size=11, color="991b1b")
    ws_dashboard[f'D{current_row}'].fill = gasto_fill
    ws_dashboard[f'D{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'D{current_row}'].border = thin_border

    ws_dashboard[f'E{current_row}'] = float(total_gastos)
    ws_dashboard[f'E{current_row}'].font = Font(bold=True, size=16, color="991b1b")
    ws_dashboard[f'E{current_row}'].fill = gasto_fill
    ws_dashboard[f'E{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'E{current_row}'].number_format = '$#,##0.00'
    ws_dashboard[f'E{current_row}'].border = thin_border

    current_row += 2

    # Balance
    balance_color = "065f46" if balance >= 0 else "991b1b"
    balance_bg = PatternFill(start_color="d1fae5" if balance >= 0 else "fee2e2",
                              end_color="d1fae5" if balance >= 0 else "fee2e2",
                              fill_type="solid")

    ws_dashboard[f'A{current_row}'] = "BALANCE"
    ws_dashboard[f'A{current_row}'].font = Font(bold=True, size=12, color=balance_color)
    ws_dashboard[f'A{current_row}'].fill = balance_bg
    ws_dashboard[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'A{current_row}'].border = thin_border

    ws_dashboard.merge_cells(f'B{current_row}:C{current_row}')
    ws_dashboard[f'B{current_row}'] = float(balance)
    ws_dashboard[f'B{current_row}'].font = Font(bold=True, size=20, color=balance_color)
    ws_dashboard[f'B{current_row}'].fill = balance_bg
    ws_dashboard[f'B{current_row}'].alignment = Alignment(horizontal='center')
    ws_dashboard[f'B{current_row}'].number_format = '$#,##0.00'
    ws_dashboard[f'B{current_row}'].border = thin_border

    current_row += 3

    # ============================================
    # TOP CATEGORÍAS DE GASTO
    # ============================================
    if categorias_data:
        ws_dashboard[f'A{current_row}'] = "TOP CATEGORÍAS DE GASTO"
        ws_dashboard[f'A{current_row}'].font = subtitle_font
        ws_dashboard.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Encabezados
        ws_dashboard[f'A{current_row}'] = "Categoría"
        ws_dashboard[f'B{current_row}'] = "Monto"
        ws_dashboard[f'C{current_row}'] = "% del Total"

        for col in ['A', 'B', 'C']:
            ws_dashboard[f'{col}{current_row}'].font = Font(bold=True, color="FFFFFF")
            ws_dashboard[f'{col}{current_row}'].fill = header_fill
            ws_dashboard[f'{col}{current_row}'].alignment = Alignment(horizontal='center')
            ws_dashboard[f'{col}{current_row}'].border = thin_border

        header_row = current_row
        current_row += 1
        data_start_row = current_row

        total_gastos_float = float(total_gastos) if total_gastos > 0 else 1

        for cat_data in categorias_data:
            ws_dashboard[f'A{current_row}'] = cat_data['categoria']
            ws_dashboard[f'B{current_row}'] = float(cat_data['total'])
            ws_dashboard[f'B{current_row}'].number_format = '$#,##0.00'
            ws_dashboard[f'C{current_row}'] = f"=B{current_row}/{total_gastos_float}"
            ws_dashboard[f'C{current_row}'].number_format = '0.00%'

            # Bordes
            for col in ['A', 'B', 'C']:
                ws_dashboard[f'{col}{current_row}'].border = thin_border

            current_row += 1

        data_end_row = current_row - 1

        # Crear gráfico de barras para categorías
        if data_end_row >= data_start_row:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Gastos por Categoría"
            chart.y_axis.title = 'Monto ($)'
            chart.x_axis.title = 'Categoría'
            chart.height = 10
            chart.width = 20

            data = Reference(ws_dashboard, min_col=2, min_row=header_row, max_row=data_end_row)
            cats = Reference(ws_dashboard, min_col=1, min_row=data_start_row, max_row=data_end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            ws_dashboard.add_chart(chart, f"E{header_row}")

    # Ajustar anchos de columna
    ws_dashboard.column_dimensions['A'].width = 25
    ws_dashboard.column_dimensions['B'].width = 18
    ws_dashboard.column_dimensions['C'].width = 15
    ws_dashboard.column_dimensions['D'].width = 18
    ws_dashboard.column_dimensions['E'].width = 18
    ws_dashboard.column_dimensions['F'].width = 15

    # ============================================
    # HOJA 2: DATOS DE TRANSACCIONES
    # ============================================
    ws_data = wb.create_sheet("Transacciones")

    # Encabezados
    headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Datos
    row_num = 2
    for trans in transacciones:
        ws_data.cell(row=row_num, column=1, value=trans.fecha.strftime('%d/%m/%Y'))
        ws_data.cell(row=row_num, column=2, value=trans.tipo)
        ws_data.cell(row=row_num, column=3, value=trans.categoria)
        ws_data.cell(row=row_num, column=4, value=trans.descripcion or '-')
        monto_cell = ws_data.cell(row=row_num, column=5, value=float(trans.monto))
        monto_cell.number_format = '$#,##0.00'

        # Colorear filas según tipo
        if trans.tipo == 'INGRESO':
            fill_color = ingreso_fill
        else:
            fill_color = gasto_fill

        for col in range(1, 6):
            ws_data.cell(row=row_num, column=col).fill = fill_color
            ws_data.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    # Ajustar anchos de columna
    ws_data.column_dimensions['A'].width = 12
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 20
    ws_data.column_dimensions['D'].width = 40
    ws_data.column_dimensions['E'].width = 15

    # ============================================
    # HOJA 3: GRÁFICO RESUMEN
    # ============================================
    ws_chart = wb.create_sheet("Gráficos")

    # Datos para gráfico de Ingresos vs Gastos
    ws_chart['A1'] = "Resumen Financiero"
    ws_chart['A1'].font = title_font
    ws_chart.merge_cells('A1:C1')

    ws_chart['A3'] = "Concepto"
    ws_chart['B3'] = "Monto"
    for col in ['A', 'B']:
        ws_chart[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_chart[f'{col}3'].fill = header_fill
        ws_chart[f'{col}3'].alignment = Alignment(horizontal='center')
        ws_chart[f'{col}3'].border = thin_border

    ws_chart['A4'] = "Ingresos"
    ws_chart['B4'] = float(total_ingresos)
    ws_chart['B4'].number_format = '$#,##0.00'
    ws_chart['A4'].fill = ingreso_fill
    ws_chart['B4'].fill = ingreso_fill

    ws_chart['A5'] = "Gastos"
    ws_chart['B5'] = float(total_gastos)
    ws_chart['B5'].number_format = '$#,##0.00'
    ws_chart['A5'].fill = gasto_fill
    ws_chart['B5'].fill = gasto_fill

    for row in [4, 5]:
        for col in ['A', 'B']:
            ws_chart[f'{col}{row}'].border = thin_border

    # Crear gráfico circular
    pie = PieChart()
    pie.title = "Distribución Ingresos vs Gastos"
    pie.height = 12
    pie.width = 18

    labels = Reference(ws_chart, min_col=1, min_row=4, max_row=5)
    data_pie = Reference(ws_chart, min_col=2, min_row=3, max_row=5)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(labels)

    # Colores del gráfico
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

    ws_chart.add_chart(pie, "D3")

    # Ajustar anchos
    ws_chart.column_dimensions['A'].width = 15
    ws_chart.column_dimensions['B'].width = 15

    # ============================================
    # GUARDAR Y RESPONDER
    # ============================================

    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Nombre del archivo
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="Transacciones_{request.user.username}_{fecha_actual}.xlsx"'

    # Guardar el libro en la respuesta
    wb.save(response)

    return response